import cv2
import mediapipe as mp
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from tkinter import Tk, filedialog
from collections import deque
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


WIN_NAME = "Bike Fit Inteligente (aperte 's' para sair)"


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("ASCII")


def resize_with_letterbox(frame: np.ndarray, target_w: int, target_h: int) -> np.ndarray:
    h, w = frame.shape[:2]
    if w <= 0 or h <= 0 or target_w <= 0 or target_h <= 0:
        return frame

    scale = min(target_w / w, target_h / h)
    new_w = max(1, int(w * scale))
    new_h = max(1, int(h * scale))

    resized = cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
    output = np.zeros((target_h, target_w, 3), dtype=np.uint8)

    x_off = (target_w - new_w) // 2
    y_off = (target_h - new_h) // 2
    output[y_off:y_off + new_h, x_off:x_off + new_w] = resized
    return output


def draw_text_box(
    frame,
    text,
    pos,
    font_scale=1.95,
    thickness=6,
    text_color=(255, 255, 255),
    box_color=(0, 0, 0),
    alpha=0.30,
    line_type=cv2.LINE_AA
):
    font = cv2.FONT_HERSHEY_SIMPLEX
    text = normalize_text(text)

    (w, h), _ = cv2.getTextSize(text, font, font_scale, thickness)
    x, y = pos
    padding = 6

    overlay = frame.copy()
    cv2.rectangle(
        overlay,
        (x - padding, y - h - padding),
        (x + w + padding, y + padding),
        box_color,
        -1
    )
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)

    cv2.putText(
        frame,
        text,
        (x, y),
        font,
        font_scale,
        text_color,
        thickness,
        line_type
    )


def draw_small_label(
    frame,
    text,
    pos,
    font_scale=1.15,
    thickness=3,
    text_color=(255, 255, 255),
    box_color=(0, 0, 0),
    alpha=0.85,
    line_type=cv2.LINE_AA
):
    font = cv2.FONT_HERSHEY_SIMPLEX
    text = normalize_text(text)

    (w, h), _ = cv2.getTextSize(text, font, font_scale, thickness)
    x, y = pos
    padding = 6

    overlay = frame.copy()
    cv2.rectangle(
        overlay,
        (x - padding, y - h - padding),
        (x + w + padding, y + padding),
        box_color,
        -1
    )
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)

    cv2.putText(
        frame,
        text,
        (x, y),
        font,
        font_scale,
        text_color,
        thickness,
        line_type
    )


def label_deg(value: float) -> str:
    return f"{int(round(value))} graus"


mp_drawing = mp.solutions.drawing_utils
mp_pose = mp.solutions.pose


def draw_one_side_no_face(frame, pose_landmarks, side="right", thickness=3, circle_radius=4):
    PL = mp_pose.PoseLandmark
    h, w = frame.shape[:2]

    if side.lower() == "right":
        pontos = [
            PL.RIGHT_SHOULDER,
            PL.RIGHT_ELBOW,
            PL.RIGHT_WRIST,
            PL.RIGHT_HIP,
            PL.RIGHT_KNEE,
            PL.RIGHT_ANKLE,
            PL.RIGHT_HEEL,
            PL.RIGHT_FOOT_INDEX
        ]

        conexoes = [
            (PL.RIGHT_SHOULDER, PL.RIGHT_ELBOW),
            (PL.RIGHT_ELBOW, PL.RIGHT_WRIST),
            (PL.RIGHT_SHOULDER, PL.RIGHT_HIP),
            (PL.RIGHT_HIP, PL.RIGHT_KNEE),
            (PL.RIGHT_KNEE, PL.RIGHT_ANKLE),
            (PL.RIGHT_ANKLE, PL.RIGHT_HEEL),
            (PL.RIGHT_HEEL, PL.RIGHT_FOOT_INDEX),
            (PL.RIGHT_ANKLE, PL.RIGHT_FOOT_INDEX),
        ]
    else:
        pontos = [
            PL.LEFT_SHOULDER,
            PL.LEFT_ELBOW,
            PL.LEFT_WRIST,
            PL.LEFT_HIP,
            PL.LEFT_KNEE,
            PL.LEFT_ANKLE,
            PL.LEFT_HEEL,
            PL.LEFT_FOOT_INDEX
        ]

        conexoes = [
            (PL.LEFT_SHOULDER, PL.LEFT_ELBOW),
            (PL.LEFT_ELBOW, PL.LEFT_WRIST),
            (PL.LEFT_SHOULDER, PL.LEFT_HIP),
            (PL.LEFT_HIP, PL.LEFT_KNEE),
            (PL.LEFT_KNEE, PL.LEFT_ANKLE),
            (PL.LEFT_ANKLE, PL.LEFT_HEEL),
            (PL.LEFT_HEEL, PL.LEFT_FOOT_INDEX),
            (PL.LEFT_ANKLE, PL.LEFT_FOOT_INDEX),
        ]

    lm = pose_landmarks.landmark

    for a, b in conexoes:
        p1 = lm[a.value]
        p2 = lm[b.value]

        x1, y1 = int(p1.x * w), int(p1.y * h)
        x2, y2 = int(p2.x * w), int(p2.y * h)

        cv2.line(frame, (x1, y1), (x2, y2), (255, 255, 255), thickness)

    for p in pontos:
        ponto = lm[p.value]
        x, y = int(ponto.x * w), int(ponto.y * h)

        cv2.circle(frame, (x, y), circle_radius, (255, 255, 255), -1)


def angle_between_points(a, b, c):
    a, b, c = np.array(a), np.array(b), np.array(c)
    ba = a - b
    bc = c - b
    denom = np.linalg.norm(ba) * np.linalg.norm(bc)
    if denom == 0:
        return np.nan
    cos_angle = np.clip(np.dot(ba, bc) / denom, -1.0, 1.0)
    return float(np.degrees(np.arccos(cos_angle)))


def landmark_to_xy(landmark, shape):
    h, w = shape
    return int(landmark.x * w), int(landmark.y * h)


def smooth_signal(values, window=7):
    if len(values) < window:
        return np.nan
    return float(np.nanmean(list(values)[-window:]))


def foot_inclination_deg(heel_xy, toe_xy):
    v = np.array(toe_xy, dtype=float) - np.array(heel_xy, dtype=float)
    if np.linalg.norm(v) == 0:
        return np.nan
    return float(np.degrees(np.arctan2(v[1], v[0])))


def compute_angles(landmarks, shape):
    idx = {
        "LS": 11, "RS": 12,
        "LE": 13, "RE": 14,
        "LW": 15, "RW": 16,
        "LH": 23, "RH": 24,
        "LK": 25, "RK": 26,
        "LA": 27, "RA": 28,
        "LHE": 29, "RHE": 30,
        "LFI": 31, "RFI": 32
    }
    pts = {k: landmark_to_xy(landmarks[i], shape) for k, i in idx.items()}

    angles = {}

    angles["knee_left"] = angle_between_points(pts["LH"], pts["LK"], pts["LA"])
    angles["knee_right"] = angle_between_points(pts["RH"], pts["RK"], pts["RA"])

    shoulders_mid = np.mean([pts["LS"], pts["RS"]], axis=0)
    hips_mid = np.mean([pts["LH"], pts["RH"]], axis=0)
    trunk_vec = shoulders_mid - hips_mid
    vertical = np.array([0.0, -1.0])
    denom = np.linalg.norm(trunk_vec) * np.linalg.norm(vertical)
    angles["trunk"] = np.nan if denom == 0 else float(
        np.degrees(np.arccos(np.clip(np.dot(trunk_vec, vertical) / denom, -1.0, 1.0)))
    )

    angles["elbow_left"] = angle_between_points(pts["LS"], pts["LE"], pts["LW"])
    angles["elbow_right"] = angle_between_points(pts["RS"], pts["RE"], pts["RW"])

    angles["hand_drop_left_px"] = float(pts["LW"][1] - pts["LS"][1])
    angles["hand_drop_right_px"] = float(pts["RW"][1] - pts["RS"][1])

    angles["ankle_left"] = angle_between_points(pts["LK"], pts["LA"], pts["LFI"])
    angles["ankle_right"] = angle_between_points(pts["RK"], pts["RA"], pts["RFI"])

    angles["foot_incl_left_deg"] = foot_inclination_deg(pts["LHE"], pts["LFI"])
    angles["foot_incl_right_deg"] = foot_inclination_deg(pts["RHE"], pts["RFI"])

    angles["foot_toe_minus_heel_dy_left_px"] = float(pts["LFI"][1] - pts["LHE"][1])
    angles["foot_toe_minus_heel_dy_right_px"] = float(pts["RFI"][1] - pts["RHE"][1])

    angles["hip_left"] = angle_between_points(pts["LS"], pts["LH"], pts["LK"])
    angles["hip_right"] = angle_between_points(pts["RS"], pts["RH"], pts["RK"])

    angles["shoulder_left"] = angle_between_points(pts["LE"], pts["LS"], pts["LH"])
    angles["shoulder_right"] = angle_between_points(pts["RE"], pts["RS"], pts["RH"])

    angles["pts"] = pts
    angles["hips_mid"] = tuple(np.array(hips_mid, dtype=int))
    return angles


class DebouncedStatus:
    def __init__(self, initial="sem_leitura", stable_frames=12):
        self.current = initial
        self.candidate = initial
        self.count = 0
        self.stable_frames = stable_frames

    def update(self, new_candidate: str) -> str:
        if new_candidate == self.candidate:
            self.count += 1
        else:
            self.candidate = new_candidate
            self.count = 1

        if self.count >= self.stable_frames:
            self.current = self.candidate
        return self.current


def classify_selim(knee_min):
    if np.isnan(knee_min):
        return "sem_leitura"
    if knee_min > 150:
        return "alto"
    if knee_min < 140:
        return "baixo"
    return "ok"


def classify_tronco(trunk):
    if np.isnan(trunk):
        return "sem_leitura"
    if trunk > 60:
        return "agressivo"
    if trunk < 20:
        return "muito_ereto"
    return "ok"


def classify_assim(asym):
    if np.isnan(asym):
        return "sem_leitura"
    return "sim" if asym > 5 else "nao"


def classify_elbow(elbow_angle):
    if np.isnan(elbow_angle):
        return "sem_leitura"
    if elbow_angle > 165:
        return "longe"
    if elbow_angle < 145:
        return "perto"
    return "ok"


def classify_foot(ankle_angle):
    if np.isnan(ankle_angle):
        return "sem_leitura"
    if ankle_angle < 90:
        return "ponta_baixa"
    if ankle_angle > 110:
        return "ponta_alta"
    return "ok"


def format_live_msgs(selim_status, tronco_status, assim_status, guidao_status, pe_status):
    msgs = []

    if selim_status == "baixo":
        msgs.append("Selim baixo (joelho muito flexionado no BDC)")
    elif selim_status == "alto":
        msgs.append("Selim alto (joelho muito estendido no BDC)")
    elif selim_status == "ok":
        msgs.append("Altura do selim adequada")
    else:
        msgs.append("Selim: sem leitura")

    if tronco_status == "agressivo":
        msgs.append("Postura agressiva - guidao baixo")
    elif tronco_status == "muito_ereto":
        msgs.append("Postura muito ereta - guidao alto")
    elif tronco_status == "ok":
        msgs.append("Inclinacao do tronco adequada")
    else:
        msgs.append("Tronco: sem leitura")

    if assim_status == "sim":
        msgs.append("Assimetria entre as pernas detectada")
    elif assim_status == "nao":
        msgs.append("Assimetria: OK")
    else:
        msgs.append("Assimetria: sem leitura")

    if guidao_status == "longe":
        msgs.append("Guidao: longe (cotovelo muito estendido)")
    elif guidao_status == "perto":
        msgs.append("Guidao: perto (cotovelo muito fechado)")
    elif guidao_status == "ok":
        msgs.append("Guidao: alcance adequado")
    else:
        msgs.append("Guidao: sem leitura")

    if pe_status == "ponta_baixa":
        msgs.append("Pe: muito apontado para baixo")
    elif pe_status == "ponta_alta":
        msgs.append("Pe: muito elevado")
    elif pe_status == "ok":
        msgs.append("Pe: movimento adequado")
    else:
        msgs.append("Pe: sem leitura")

    return msgs


def final_recommendations(selim_status, tronco_status, assimetria_status, guidao_status, pe_status):
    recs = []

    if selim_status == "baixo":
        recs.append("Selim: subir 2 a 5 mm e regravar.")
    elif selim_status == "alto":
        recs.append("Selim: descer 2 a 5 mm e regravar.")
    elif selim_status == "ok":
        recs.append("Selim: OK (ajuste fino 1 a 2 mm se houver desconforto).")
    else:
        recs.append("Selim: sem leitura (melhorar iluminacao e camera bem de lado).")

    if tronco_status == "agressivo":
        recs.append("Tronco/Guidao: muito inclinado. Elevar guidao (spacers/mesa) ou encurtar mesa pode ajudar.")
    elif tronco_status == "muito_ereto":
        recs.append("Tronco/Guidao: muito ereto. Baixar um pouco guidao ou usar mesa ligeiramente mais longa (com cuidado).")
    elif tronco_status == "ok":
        recs.append("Tronco/Guidao: OK.")
    else:
        recs.append("Tronco: sem leitura (camera alinhada e corpo inteiro visivel).")

    if guidao_status == "longe":
        recs.append("Reach: guidao longe. Tente mesa mais curta/aproximar o guidao e reavaliar.")
    elif guidao_status == "perto":
        recs.append("Reach: guidao perto. Tente mesa mais longa/afastar o guidao e reavaliar.")
    elif guidao_status == "ok":
        recs.append("Reach: OK.")
    else:
        recs.append("Reach: sem leitura (punho/cotovelo precisam aparecer bem).")

    if pe_status == "ponta_baixa":
        recs.append("Pe/Tornozelo: muita flexao plantar (ponta baixa). Verifique tecnica e posicao do taco.")
    elif pe_status == "ponta_alta":
        recs.append("Pe/Tornozelo: muita dorsiflexao (ponta alta). Ajuste tecnica e confira altura do selim/alcance.")
    elif pe_status == "ok":
        recs.append("Pe/Tornozelo: OK.")
    else:
        recs.append("Pe/Tornozelo: sem leitura (ponta e calcanhar precisam aparecer bem).")

    if assimetria_status == "sim":
        recs.append("Assimetria: confirme camera 100% lateral. Se persistir, revise tacos/pe (cleat) e centralizacao do selim.")
    elif assimetria_status == "nao":
        recs.append("Assimetria: OK.")
    else:
        recs.append("Assimetria: sem leitura.")

    return " | ".join(recs)


def formatar_excel_cliente(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    ws.title = "Resumo Bike Fit"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    fill_ok = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fill_alerta = PatternFill(fill_type="solid", fgColor="FFF2CC")
    fill_ajuste = PatternFill(fill_type="solid", fgColor="F4CCCC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if idx in [2, 3]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

        situacao_cell = row[2]
        situacao = str(situacao_cell.value).strip().lower()

        if situacao == "adequado":
            situacao_cell.fill = fill_ok
        elif situacao == "atenção":
            situacao_cell.fill = fill_alerta
        elif situacao == "ajuste necessário":
            situacao_cell.fill = fill_ajuste

    larguras = {
        "A": 24,
        "B": 12,
        "C": 22,
        "D": 45
    }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.row_dimensions[1].height = 24
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 32

    ws.freeze_panes = "A2"
    wb.save(caminho_arquivo)


def gerar_tabela_cliente_excel(resumo_df, output_excel="bikefit_tabela_cliente.xlsx"):
    linha = resumo_df.iloc[0]
    tabela = []

    def adicionar_parametro(parametro, valor, situacao, recomendacao):
        tabela.append({
            "Parâmetro": parametro,
            "Valor": valor,
            "Situação": situacao,
            "Recomendação": recomendacao
        })

    joelho = float(linha["Joelho médio (graus)"])
    status_selim = str(linha["Status do selim"])
    if status_selim == "baixo":
        adicionar_parametro("Joelho / Selim", f"{joelho:.1f}°", "Ajuste necessário", "Subir o selim entre 2 e 5 mm.")
    elif status_selim == "alto":
        adicionar_parametro("Joelho / Selim", f"{joelho:.1f}°", "Ajuste necessário", "Descer o selim entre 2 e 5 mm.")
    else:
        adicionar_parametro("Joelho / Selim", f"{joelho:.1f}°", "Adequado", "Manter ajuste atual.")

    tronco = float(linha["Tronco médio (graus)"])
    status_tronco = str(linha["Status do tronco"])
    if status_tronco == "agressivo":
        adicionar_parametro("Tronco / Guidão", f"{tronco:.1f}°", "Ajuste necessário", "Elevar o guidão ou encurtar a frente.")
    elif status_tronco == "muito_ereto":
        adicionar_parametro("Tronco / Guidão", f"{tronco:.1f}°", "Ajuste necessário", "Avaliar redução leve da altura do guidão.")
    else:
        adicionar_parametro("Tronco / Guidão", f"{tronco:.1f}°", "Adequado", "Manter ajuste atual.")

    cotovelo = float(linha["Cotovelo médio (graus)"])
    status_guidao = str(linha["Status do guidão"])
    if status_guidao == "longe":
        adicionar_parametro("Cotovelo / Alcance", f"{cotovelo:.1f}°", "Ajuste necessário", "Aproximar o guidão ou usar mesa mais curta.")
    elif status_guidao == "perto":
        adicionar_parametro("Cotovelo / Alcance", f"{cotovelo:.1f}°", "Ajuste necessário", "Afastar o guidão ou usar mesa mais longa.")
    else:
        adicionar_parametro("Cotovelo / Alcance", f"{cotovelo:.1f}°", "Adequado", "Manter ajuste atual.")

    tornozelo = float(linha["Tornozelo médio (graus)"])
    status_pe = str(linha["Status do pé"])
    if status_pe == "ponta_baixa":
        adicionar_parametro("Tornozelo / Pé", f"{tornozelo:.1f}°", "Atenção", "Revisar técnica de pedalada e posição do pé.")
    elif status_pe == "ponta_alta":
        adicionar_parametro("Tornozelo / Pé", f"{tornozelo:.1f}°", "Atenção", "Revisar técnica, altura do selim e posição do pé.")
    else:
        adicionar_parametro("Tornozelo / Pé", f"{tornozelo:.1f}°", "Adequado", "Manter padrão atual.")

    assimetria = float(linha["Assimetria média (graus)"])
    status_assim = str(linha["Status da assimetria"])
    if status_assim == "sim":
        adicionar_parametro("Assimetria", f"{assimetria:.1f}°", "Atenção", "Revisar alinhamento, selim e posição dos pés.")
    else:
        adicionar_parametro("Assimetria", f"{assimetria:.1f}°", "Adequado", "Manter ajuste atual e acompanhar.")

    tabela_df = pd.DataFrame(tabela)
    tabela_df.to_excel(output_excel, index=False)
    formatar_excel_cliente(output_excel)

    print("Tabela em Excel para o cliente gerada:")
    print("-", output_excel)

    return tabela_df


def gerar_graficos(df, resumo_df, pasta_saida="graficos_bikefit"):
    os.makedirs(pasta_saida, exist_ok=True)

    df_plot = df.copy()
    df_plot = df_plot.dropna(subset=["tempo_s"])

    def estilo_base():
        plt.figure(figsize=(12, 6))
        plt.grid(True, linestyle="--", alpha=0.35)

    def caixa_texto(texto):
        plt.text(
            0.02, 0.98,
            texto,
            transform=plt.gca().transAxes,
            fontsize=10,
            verticalalignment="top",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.85)
        )

    estilo_base()
    y = df_plot["joelho"]
    media = y.mean()
    plt.plot(df_plot["tempo_s"], y, linewidth=2, label="Joelho")
    plt.axhspan(140, 150, alpha=0.15, label="Faixa ideal (140°–150°)")
    plt.axhline(media, linestyle="--", linewidth=2, label=f"Média = {media:.1f}°")
    plt.xlabel("Tempo (s)")
    plt.ylabel("Ângulo do joelho (graus)")
    plt.title("Variação do ângulo do joelho ao longo do tempo")
    caixa_texto(
        "Este gráfico mostra a variação do joelho durante a pedalada.\n"
        "A faixa sombreada representa a zona recomendada no ponto mais baixo do pedal.\n"
        "Valores abaixo sugerem selim baixo; acima, selim alto."
    )
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "joelho_tempo.png"), dpi=300, bbox_inches="tight")
    plt.close()

    estilo_base()
    y = df_plot["tronco"]
    media = y.mean()
    plt.plot(df_plot["tempo_s"], y, linewidth=2, label="Tronco")
    plt.axhline(media, linestyle="--", linewidth=2, label=f"Média = {media:.1f}°")
    plt.xlabel("Tempo (s)")
    plt.ylabel("Inclinação do tronco (graus)")
    plt.title("Variação da inclinação do tronco ao longo do tempo")
    caixa_texto(
        "Este gráfico representa a inclinação do tronco durante o movimento.\n"
        "Valores mais altos indicam postura mais agressiva.\n"
        "Valores mais baixos indicam postura mais ereta."
    )
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "tronco_tempo.png"), dpi=300, bbox_inches="tight")
    plt.close()

    estilo_base()
    y = df_plot["cotovelo"]
    media = y.mean()
    plt.plot(df_plot["tempo_s"], y, linewidth=2, label="Cotovelo")
    plt.axhline(145, linestyle=":", linewidth=1.8, label="Limite inferior")
    plt.axhline(165, linestyle=":", linewidth=1.8, label="Limite superior")
    plt.axhline(media, linestyle="--", linewidth=2, label=f"Média = {media:.1f}°")
    plt.xlabel("Tempo (s)")
    plt.ylabel("Ângulo do cotovelo (graus)")
    plt.title("Variação do ângulo do cotovelo ao longo do tempo")
    caixa_texto(
        "Este gráfico ajuda a avaliar o alcance do guidão.\n"
        "Ângulos muito abertos podem indicar guidão distante.\n"
        "Ângulos muito fechados podem indicar guidão próximo."
    )
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "cotovelo_tempo.png"), dpi=300, bbox_inches="tight")
    plt.close()

    estilo_base()
    y = df_plot["tornozelo"]
    media = y.mean()
    plt.plot(df_plot["tempo_s"], y, linewidth=2, label="Tornozelo")
    plt.axhline(90, linestyle=":", linewidth=1.8, label="Referência inferior")
    plt.axhline(110, linestyle=":", linewidth=1.8, label="Referência superior")
    plt.axhline(media, linestyle="--", linewidth=2, label=f"Média = {media:.1f}°")
    plt.xlabel("Tempo (s)")
    plt.ylabel("Ângulo do tornozelo (graus)")
    plt.title("Variação do ângulo do tornozelo ao longo do tempo")
    plt.text(
        0.02, 0.05,
        "Este gráfico mostra o comportamento do tornozelo na pedalada.\n"
        "Valores baixos indicam ponta do pé para baixo.\n"
        "Valores altos indicam o pé mais elevado.",
        transform=plt.gca().transAxes,
        fontsize=10,
        verticalalignment="bottom",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.85)
    )
    plt.legend(loc="upper right")
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "tornozelo_tempo.png"), dpi=300, bbox_inches="tight")
    plt.close()

    estilo_base()
    y = df_plot["assimetria"]
    media = y.mean()
    plt.plot(df_plot["tempo_s"], y, linewidth=2, label="Assimetria")
    plt.axhspan(0, 5, alpha=0.15, label="Faixa aceitável (até 5°)")
    plt.axhline(media, linestyle="--", linewidth=2, label=f"Média = {media:.1f}°")
    plt.axhline(5, linestyle=":", linewidth=1.8, label="Limite de atenção")
    plt.xlabel("Tempo (s)")
    plt.ylabel("Diferença angular (graus)")
    plt.title("Variação da assimetria entre as pernas ao longo do tempo")
    caixa_texto(
        "Este gráfico indica a diferença angular entre os lados do corpo.\n"
        "Valores acima de 5° podem sugerir assimetria postural\n"
        "ou necessidade de revisar posicionamento e ajuste."
    )
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "assimetria_tempo.png"), dpi=300, bbox_inches="tight")
    plt.close()

    medias = {
        "Joelho": float(resumo_df.loc[0, "Joelho médio (graus)"]),
        "Tronco": float(resumo_df.loc[0, "Tronco médio (graus)"]),
        "Cotovelo": float(resumo_df.loc[0, "Cotovelo médio (graus)"]),
        "Quadril": float(resumo_df.loc[0, "Quadril médio (graus)"]),
        "Ombro": float(resumo_df.loc[0, "Ombro médio (graus)"]),
        "Tornozelo": float(resumo_df.loc[0, "Tornozelo médio (graus)"]),
    }

    estilo_base()
    barras = plt.bar(list(medias.keys()), list(medias.values()))
    for barra, valor in zip(barras, medias.values()):
        plt.text(
            barra.get_x() + barra.get_width() / 2,
            barra.get_height() + 1,
            f"{valor:.1f}°",
            ha="center",
            va="bottom",
            fontsize=10
        )
    plt.xlabel("Parâmetros biomecânicos")
    plt.ylabel("Ângulo médio (graus)")
    plt.title("Comparação das médias angulares finais")
    caixa_texto(
        "Este gráfico resume os valores médios calculados durante a análise.\n"
        "Ele permite comparar os principais parâmetros biomecânicos\n"
        "utilizados para interpretar a postura do ciclista."
    )
    plt.tight_layout()
    plt.savefig(os.path.join(pasta_saida, "medias_finais.png"), dpi=300, bbox_inches="tight")
    plt.close()

    print(f"Gráficos salvos na pasta: {pasta_saida}")


def melhorar_qualidade_frame(frame):
    frame = cv2.convertScaleAbs(frame, alpha=1.08, beta=4)

    blur = cv2.GaussianBlur(frame, (0, 0), 1.0)
    frame = cv2.addWeighted(frame, 1.08, blur, -0.08, 0)
    return frame


def run_bikefit(
    video_path="IMG_5683.mov",
    output_csv="bikefit_resultado_final.csv",
    output_excel_cliente="bikefit_tabela_cliente.xlsx",
    draw_side="right",
    thickness=6,
    circle_radius=7
):
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise IOError("Nao foi possivel abrir o video.")

    fps_video = cap.get(cv2.CAP_PROP_FPS)
    if fps_video <= 0:
        fps_video = 30.0

    # =========================================
    # AJUSTE AUTOMÁTICO DE ESCALA
    # =========================================

    video_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    video_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    escala_video = min(video_w / 1920, video_h / 1080)
    if video_w < 1000:
        escala_video *= 0.55

    if video_w < 1000:
        font_scale_main = 0.38
        font_scale_small = 0.28

        pos_x_texto = 12
        pos_y_texto = 28
        alpha_box = 0.72
    else:
        font_scale_main = 1.95
        font_scale_small = 1.15

        pos_x_texto = 25
        pos_y_texto = 75
        alpha_box = 0.93

    espaco_texto = 18 if video_w < 1000 else max(32, int(70 * escala_video))

    if video_w < 1000:
        thickness_text = 1
        thickness_small = 1
    else:
        thickness_text = max(1, int(4 * escala_video))
        thickness_small = max(1, int(2 * escala_video))

    # =========================================
    # AJUSTE AUTOMÁTICO DE NITIDEZ DAS LETRAS
    # =========================================

    line_type_text = cv2.LINE_AA

    cv2.namedWindow(WIN_NAME, cv2.WINDOW_NORMAL)

    knee_history = deque(maxlen=120)
    trunk_history = deque(maxlen=30)
    elbow_history = deque(maxlen=30)
    ankle_history = deque(maxlen=30)
    foot_incl_history = deque(maxlen=30)
    hip_history = deque(maxlen=30)
    shoulder_history = deque(maxlen=30)

    selim_db = DebouncedStatus(initial="sem_leitura", stable_frames=12)
    tronco_db = DebouncedStatus(initial="sem_leitura", stable_frames=12)
    assim_db = DebouncedStatus(initial="sem_leitura", stable_frames=12)
    guidao_db = DebouncedStatus(initial="sem_leitura", stable_frames=12)
    pe_db = DebouncedStatus(initial="sem_leitura", stable_frames=12)

    records = []
    frame_idx = 0
    side = "right" if str(draw_side).lower().startswith("r") else "left"

    with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
        while True:
            ret, frame = cap.read()
            if not ret:
                break

            # =========================================
            # MELHORA DE RESOLUÇÃO PARA VÍDEOS PEQUENOS
            # =========================================

            if video_w < 1000:
                frame = cv2.resize(
                    frame,
                    None,
                    fx=1.8,
                    fy=1.8,
                    interpolation=cv2.INTER_CUBIC
                )

            frame = melhorar_qualidade_frame(frame)
            frame_idx += 1
            timestamp = frame_idx / fps_video

            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            result = pose.process(rgb)

            knee = elbow = trunk = asym = np.nan
            ankle_ang = foot_incl = toe_minus_heel_dy = np.nan
            hip_ang = shoulder_ang = np.nan

            sx = sy = ex = ey = hxip = hyip = kx = ky = ax = ay = tx = ty = np.nan

            if result.pose_landmarks:
                angles = compute_angles(result.pose_landmarks.landmark, frame.shape[:2])
                pts = angles["pts"]
                trunk = angles["trunk"]

                if side == "right":
                    knee = angles["knee_right"]
                    elbow = angles["elbow_right"]
                    ankle_ang = angles["ankle_right"]
                    foot_incl = angles["foot_incl_right_deg"]
                    toe_minus_heel_dy = angles["foot_toe_minus_heel_dy_right_px"]
                    hip_ang = angles["hip_right"]
                    shoulder_ang = angles["shoulder_right"]

                    sx, sy = pts["RS"]
                    ex, ey = pts["RE"]
                    hxip, hyip = pts["RH"]
                    kx, ky = pts["RK"]
                    ax, ay = pts["RA"]
                    tx, ty = pts["RFI"]
                else:
                    knee = angles["knee_left"]
                    elbow = angles["elbow_left"]
                    ankle_ang = angles["ankle_left"]
                    foot_incl = angles["foot_incl_left_deg"]
                    toe_minus_heel_dy = angles["foot_toe_minus_heel_dy_left_px"]
                    hip_ang = angles["hip_left"]
                    shoulder_ang = angles["shoulder_left"]

                    sx, sy = pts["LS"]
                    ex, ey = pts["LE"]
                    hxip, hyip = pts["LH"]
                    kx, ky = pts["LK"]
                    ax, ay = pts["LA"]
                    tx, ty = pts["LFI"]

                knee_history.append(knee)
                trunk_history.append(trunk)
                elbow_history.append(elbow)
                ankle_history.append(ankle_ang)
                foot_incl_history.append(foot_incl)
                hip_history.append(hip_ang)
                shoulder_history.append(shoulder_ang)

                knee_min = float(np.nanmin(knee_history)) if len(knee_history) else np.nan
                trunk_smooth = smooth_signal(trunk_history)
                elbow_smooth = smooth_signal(elbow_history)
                ankle_smooth = smooth_signal(ankle_history)
                foot_incl_smooth = smooth_signal(foot_incl_history)
                hip_smooth = smooth_signal(hip_history)
                shoulder_smooth = smooth_signal(shoulder_history)

                asym = float(abs(angles["knee_left"] - angles["knee_right"]))

                selim_status = selim_db.update(classify_selim(knee_min))
                tronco_status = tronco_db.update(classify_tronco(trunk_smooth))
                assim_status = assim_db.update(classify_assim(asym))
                guidao_status = guidao_db.update(classify_elbow(elbow_smooth))
                pe_status = pe_db.update(classify_foot(ankle_smooth))

                draw_one_side_no_face(
                    frame,
                    result.pose_landmarks,
                    side=side,
                    thickness=thickness,
                    circle_radius=circle_radius
                )

                msgs = format_live_msgs(
                    selim_status,
                    tronco_status,
                    assim_status,
                    guidao_status,
                    pe_status
                )

                for i, msg in enumerate(msgs):
                    draw_text_box(
                        frame,
                        msg,
                        pos=(pos_x_texto, pos_y_texto + i * espaco_texto),
                        font_scale=font_scale_main,
                        thickness=thickness_text,
                        alpha=alpha_box,
                        line_type=line_type_text
                    )

                if not np.isnan(shoulder_smooth) and not np.isnan(sx) and not np.isnan(sy):
                    draw_small_label(
                        frame,
                        label_deg(shoulder_smooth),
                        pos=(int(sx) + 12, int(sy) - 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(elbow_smooth) and not np.isnan(ex) and not np.isnan(ey):
                    draw_small_label(
                        frame,
                        label_deg(elbow_smooth),
                        pos=(int(ex) + 12, int(ey) - 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(hip_smooth) and not np.isnan(hxip) and not np.isnan(hyip):
                    draw_small_label(
                        frame,
                        label_deg(hip_smooth),
                        pos=(int(hxip) + 12, int(hyip) - 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(knee) and not np.isnan(kx) and not np.isnan(ky):
                    draw_small_label(
                        frame,
                        label_deg(knee),
                        pos=(int(kx) + 12, int(ky) - 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(ankle_smooth) and not np.isnan(ax) and not np.isnan(ay):
                    draw_small_label(
                        frame,
                        label_deg(ankle_smooth),
                        pos=(int(ax) + 12, int(ay) - 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(foot_incl_smooth) and not np.isnan(tx) and not np.isnan(ty):
                    draw_small_label(
                        frame,
                        label_deg(foot_incl_smooth),
                        pos=(int(tx) + 12, int(ty) + 12),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                if not np.isnan(trunk_smooth) and not np.isnan(hxip) and not np.isnan(hyip):
                    draw_small_label(
                        frame,
                        f"T{label_deg(trunk_smooth)}",
                        pos=(int(hxip) + 12, int(hyip) - 42),
                        font_scale=font_scale_small,
                        thickness=thickness_small,
                        line_type=line_type_text
                    )

                records.append({
                    "tempo_s": timestamp,
                    "lado_usado": side,
                    "joelho": float(knee) if not np.isnan(knee) else np.nan,
                    "tronco": float(trunk_smooth) if not np.isnan(trunk_smooth) else np.nan,
                    "assimetria": float(asym) if not np.isnan(asym) else np.nan,
                    "cotovelo": float(elbow_smooth) if not np.isnan(elbow_smooth) else np.nan,
                    "quadril": float(hip_smooth) if not np.isnan(hip_smooth) else np.nan,
                    "ombro": float(shoulder_smooth) if not np.isnan(shoulder_smooth) else np.nan,
                    "tornozelo": float(ankle_smooth) if not np.isnan(ankle_smooth) else np.nan,
                    "pe_inclinacao_deg": float(foot_incl_smooth) if not np.isnan(foot_incl_smooth) else np.nan,
                    "ponta_menos_calcanhar_dy_px": float(toe_minus_heel_dy) if not np.isnan(toe_minus_heel_dy) else np.nan,
                    "selim_status_live": selim_status,
                    "tronco_status_live": tronco_status,
                    "assimetria_status_live": assim_status,
                    "guidao_status_live": guidao_status,
                    "pe_status_live": pe_status,
                })

            x, y, w, h = cv2.getWindowImageRect(WIN_NAME)
            frame_show = resize_with_letterbox(frame, w, h) if w > 0 and h > 0 else frame
            cv2.imshow(WIN_NAME, frame_show)

            espera = 10

            if video_w < 1000:
                espera = 35

            if cv2.waitKey(espera) & 0xFF == ord("s"):
                break

    cap.release()
    cv2.destroyAllWindows()

    df = pd.DataFrame(records)
    if len(df) == 0:
        print("Sem dados suficientes para salvar.")
        return

    df = df.round({
        "tempo_s": 2,
        "joelho": 2,
        "tronco": 2,
        "assimetria": 2,
        "cotovelo": 2,
        "quadril": 2,
        "ombro": 2,
        "tornozelo": 2,
        "pe_inclinacao_deg": 2,
        "ponta_menos_calcanhar_dy_px": 0,
    })

    joelho_min = float(df["joelho"].min())
    joelho_medio = float(df["joelho"].mean())
    joelho_max = float(df["joelho"].max())

    tronco_medio = float(df["tronco"].mean())
    assimetria_media = float(df["assimetria"].mean())
    assimetria_max = float(df["assimetria"].max())

    cotovelo_medio = float(df["cotovelo"].mean())
    cotovelo_min = float(df["cotovelo"].min())
    cotovelo_max = float(df["cotovelo"].max())

    quadril_medio = float(df["quadril"].mean())
    ombro_medio = float(df["ombro"].mean())

    tornozelo_medio = float(df["tornozelo"].mean())
    tornozelo_min = float(df["tornozelo"].min())
    tornozelo_max = float(df["tornozelo"].max())

    pe_inclinacao_media = float(df["pe_inclinacao_deg"].mean())
    dy_pe_medio = float(df["ponta_menos_calcanhar_dy_px"].mean())

    selim_status_final = classify_selim(joelho_min)
    tronco_status_final = classify_tronco(tronco_medio)
    assimetria_status_final = classify_assim(assimetria_media)
    guidao_status_final = classify_elbow(cotovelo_medio)
    pe_status_final = classify_foot(tornozelo_medio)

    recomendacoes = final_recommendations(
        selim_status_final,
        tronco_status_final,
        assimetria_status_final,
        guidao_status_final,
        pe_status_final
    )

    resumo = {
        "Lado usado": side,
        "Joelho mínimo (graus)": joelho_min,
        "Joelho médio (graus)": joelho_medio,
        "Joelho máximo (graus)": joelho_max,
        "Tronco médio (graus)": tronco_medio,
        "Assimetria média (graus)": assimetria_media,
        "Assimetria máxima (graus)": assimetria_max,
        "Cotovelo médio (graus)": cotovelo_medio,
        "Cotovelo mínimo (graus)": cotovelo_min,
        "Cotovelo máximo (graus)": cotovelo_max,
        "Quadril médio (graus)": quadril_medio,
        "Ombro médio (graus)": ombro_medio,
        "Tornozelo médio (graus)": tornozelo_medio,
        "Tornozelo mínimo (graus)": tornozelo_min,
        "Tornozelo máximo (graus)": tornozelo_max,
        "Inclinação média do pé (graus)": pe_inclinacao_media,
        "Diferença média ponta-calcanhar (px)": dy_pe_medio,
        "Status do selim": selim_status_final,
        "Status do tronco": tronco_status_final,
        "Status da assimetria": assimetria_status_final,
        "Status do guidão": guidao_status_final,
        "Status do pé": pe_status_final,
        "Recomendações": recomendacoes
    }

    resumo_df = pd.DataFrame([resumo]).round(2)

    resumo_df.to_csv(output_csv, index=False, sep=";", decimal=",")
    gerar_tabela_cliente_excel(resumo_df, output_excel=output_excel_cliente)
    gerar_graficos(df, resumo_df)

    print("Arquivos gerados:")
    print("-", output_csv)
    print("-", output_excel_cliente)


if __name__ == "__main__":

    # =========================================
    # JANELA PARA SELECIONAR O VÍDEO
    # =========================================

    root = Tk()
    root.withdraw()

    video_escolhido = filedialog.askopenfilename(
        title="Selecione o vídeo da análise",
        filetypes=[
            ("Vídeos", "*.mp4 *.mov *.avi *.mkv"),
            ("Todos os arquivos", "*.*")
        ]
    )

    if not video_escolhido:
        print("Nenhum vídeo selecionado.")
    else:
        run_bikefit(
            video_path=video_escolhido,
            draw_side="right",
            thickness=2,
            circle_radius=3
        )
